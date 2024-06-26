from openpyxl import load_workbook
import pandas as pd

wb = load_workbook("FCHotel_FSmodeling_assumption.xlsx")
ws = wb['Sheet1']

assumption = {}
for row in ws.iter_rows(values_only=True):
    if row[0] is not None:
        assumption[row[0]] = {}
        astndct1 = assumption[row[0]]
    elif row[1] is not None:
        astndct1[row[1]] = {}
        astndct2 = astndct1[row[1]]
    elif row[2] is not None:
        astndct2[row[2]] = row[3]

monthly_occ_rate = pd.DataFrame(assumption['monthly_occ_rate'])
room_operating_cost = pd.DataFrame(assumption['room_operating_cost'])
salary_cost = pd.DataFrame(assumption['salary_cost'])